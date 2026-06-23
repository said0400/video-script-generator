"""
📋 Script Reader v2.0 — Final Production Edition

Features:
  ✅ Auto lang + content_mode detection from filename
     (ar_short.xlsx, fr_long.xlsx, en_short.xlsx, ...)
  ✅ Multi-format support (Excel, CSV)
  ✅ Auto column detection with aliases
  ✅ Positional fallback with warnings
  ✅ Tag-aware content processing (Short + Long)
  ✅ Long content: each [tag]+paragraph = one unit
  ✅ Short content: sentence-level splitting
  ✅ Duplicate number detection + auto-rename
  ✅ Validation with detailed errors
  ✅ Pretty summary display
  ✅ Robust error handling
  ✅ Word-boundary filename detection (no "archive" → "ar")
  ✅ Smart sheet selection for Excel
  ✅ CSV uses _process_rows (no code duplication)
  ✅ DEFAULT_TAG instead of None for missing tags
"""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path
from typing import Optional

from tags_parser import (
    DEFAULT_TAG,
    auto_correct_tag,
    split_into_tagged_sentences,
    strip_tags_from_text,
)

log = logging.getLogger(__name__)

# ═════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═════════════════════════════════════════════════════════════════════════════

EXCEL_EXTENSIONS = (".xlsx", ".xls")
CSV_EXTENSIONS   = (".csv",)

# Regex patterns
_NORMALIZE_RE = re.compile(r"[\s\-\/\\]")
_TAG_RE       = re.compile(r"\[([a-zA-Z_]+)\]")
_TAG_SPLIT_RE = re.compile(r"(\[[a-zA-Z_]+\])")
_FILE_LANG_RE = re.compile(
    r"(?P<lang>ar|fr|en)[_\-](?P<mode>short|long)",
    re.IGNORECASE,
)

# Sentence splitting
_AR_SENTENCE_RE = re.compile(r"(?<=[.!?؟\u06D4])\s+|\n+")
_EN_SENTENCE_RE = re.compile(r"(?<=[.!?])\s+|\n+")

# Validation
MIN_SENTENCE_CHARS = 4
MIN_SENTENCE_WORDS = 2

# Display
PREVIEW_LENGTH   = 58
TITLE_MAX_LENGTH = 50
SUMMARY_WIDTH    = 65

# Invalid number values
_INVALID_NUMBERS = frozenset({
    "number", "num", "رقم", "#", "id", "",
    "no", "n", "video_number",
})

# Supported values
_VALID_LANGS = frozenset({"ar", "fr", "en"})
_VALID_MODES = frozenset({"short", "long"})


# ═════════════════════════════════════════════════════════════════════════════
# COLUMN ALIASES
# ═════════════════════════════════════════════════════════════════════════════

COLUMN_ALIASES: dict[str, list[str]] = {
    "number": [
        "number", "num", "no", "id",
        "video_number", "رقم", "#", "n",
    ],
    "title": [
        "title", "name", "video_title", "subject",
        "عنوان", "titre", "nom",
    ],
    "ar_content": [
        "ar_content", "arabic", "ar", "arabic_content",
        "عربي", "محتوى_عربي", "النص_العربي", "content_ar",
        "نص", "محتوى",
    ],
    "en_content": [
        "en_content", "english", "en", "english_content",
        "content_en", "text_en", "script_en",
    ],
    "fr_content": [
        "fr_content", "french", "fr", "french_content",
        "contenu", "contenu_fr", "texte", "texte_fr",
        "content_fr",
    ],
    "content": [
        "content", "text", "script",
        "محتوى", "النص", "texte", "contenu",
    ],
}

LANG_CONTENT_KEYS = ("ar_content", "en_content", "fr_content")
ALL_CONTENT_KEYS  = LANG_CONTENT_KEYS + ("content",)


# ═════════════════════════════════════════════════════════════════════════════
# FILENAME METADATA DETECTION
# ═════════════════════════════════════════════════════════════════════════════

def _detect_file_meta(path: Path) -> tuple[str, str]:
    """
    استخراج lang و content_mode من اسم الملف.

    Uses word-boundary detection to avoid false matches
    (e.g., "archive" → "ar" was a bug).

    Examples:
        ar_short.xlsx      → ("ar", "short")
        videos_fr_long.xlsx → ("fr", "long")
        en_short.csv       → ("en", "short")
        my_content.xlsx    → ("ar", "short")  ← default
    """
    stem  = path.stem.lower().strip()
    match = _FILE_LANG_RE.search(stem)

    if match:
        lang = match.group("lang").lower()
        mode = match.group("mode").lower()
        log.info(
            "  📁 Detected from filename: lang=%s, mode=%s",
            lang.upper(), mode.upper()
        )
        return lang, mode

    # Fallback: word-boundary detection
    detected_lang = "ar"
    detected_mode = "short"

    # Split by common separators
    parts = re.split(r"[_\-\.\s]", stem)

    for lang in ("ar", "fr", "en"):
        # Check exact word match (not "archive" → "ar")
        if lang in parts:
            detected_lang = lang
            break

    if "long" in parts:
        detected_mode = "long"

    log.warning(
        "  ⚠️  Cannot detect lang/mode from '%s' — using %s/%s",
        path.name,
        detected_lang.upper(),
        detected_mode.upper()
    )
    return detected_lang, detected_mode


# ═════════════════════════════════════════════════════════════════════════════
# COLUMN DETECTION
# ═════════════════════════════════════════════════════════════════════════════

def _normalize(text: str) -> str:
    """تطبيع اسم العمود للمقارنة."""
    return _NORMALIZE_RE.sub("_", str(text).strip().lower())


def _find_column_index(
    normed_headers: list[str],
    aliases:        list[str],
) -> Optional[int]:
    """البحث عن index لعمود معين."""
    normed_aliases = {_normalize(a) for a in aliases}
    for i, header in enumerate(normed_headers):
        if header in normed_aliases:
            return i
    return None


def _detect_columns(
    headers:   list[str],
    file_lang: str = "ar",
) -> dict[str, int]:
    """
    اكتشاف أعمدة الملف تلقائياً.

    file_lang: يُستخدم لتحديد عمود المحتوى الافتراضي
    """
    col_map        : dict[str, int] = {}
    normed_headers = [_normalize(h) for h in headers]

    # مطابقة بالـ aliases
    for field, aliases in COLUMN_ALIASES.items():
        idx = _find_column_index(normed_headers, aliases)
        if idx is not None:
            col_map[field] = idx

    # Positional fallback مع تحذير
    if "number" not in col_map and len(headers) >= 1:
        col_map["number"] = 0
        log.warning(
            "  ⚠️  'number' column not found — using col 0 (%r)",
            headers[0]
        )
    if "title" not in col_map and len(headers) >= 2:
        col_map["title"] = 1
        log.warning(
            "  ⚠️  'title' column not found — using col 1 (%r)",
            headers[1]
        )

    # Content fallback
    if "content" in col_map:
        lang_key = f"{file_lang}_content"
        if lang_key not in col_map:
            col_map[lang_key] = col_map["content"]
            log.info(
                "  ℹ️  Using 'content' column as %s",
                lang_key
            )

    return col_map


# ═════════════════════════════════════════════════════════════════════════════
# ROW PARSING
# ═════════════════════════════════════════════════════════════════════════════

def _safe_get(row: list, idx: int) -> str:
    """قراءة آمنة لقيمة من row."""
    try:
        value = row[idx]
        if value is None:
            return ""
        return str(value).strip()
    except (IndexError, TypeError):
        return ""


def _row_to_dict(
    row:     list,
    col_map: dict[str, int],
) -> dict:
    """تحويل row إلى dict."""
    return {
        field: (
            _safe_get(row, col_map[field])
            if field in col_map else ""
        )
        for field in COLUMN_ALIASES
    }


def _has_content(record: dict) -> bool:
    """التحقق من وجود محتوى."""
    return any(
        record.get(key, "").strip()
        for key in ALL_CONTENT_KEYS
    )


def _has_valid_number(record: dict) -> bool:
    """التحقق من صحة الرقم."""
    number = record.get("number", "")
    return _normalize(str(number)) not in _INVALID_NUMBERS


def _is_valid_record(record: dict) -> bool:
    """التحقق الشامل من صحة السجل."""
    return (
        bool(record.get("title", "").strip()) and
        _has_content(record) and
        _has_valid_number(record)
    )


def _row_has_data(row: object) -> bool:
    """التحقق إذا الصف يحتوي بيانات."""
    if not row:
        return False
    try:
        return any(
            cell is not None and str(cell).strip()
            for cell in row
        )
    except Exception:
        return False


# ═════════════════════════════════════════════════════════════════════════════
# TAG-AWARE CONTENT PROCESSING
# ═════════════════════════════════════════════════════════════════════════════

def _process_single_sentence(sent: dict) -> dict:
    """
    معالجة جملة/فقرة واحدة مع tag.

    Uses DEFAULT_TAG instead of None for missing tags.
    """
    raw_tag = (sent.get("raw_tag") or "").strip()
    text    = (sent.get("text")    or "").strip()
    line    = sent.get("line", 1)

    # تصحيح الـ tag
    if raw_tag:
        corrected, reason = auto_correct_tag(raw_tag)
        if corrected:
            final_tag  = corrected
            tag_source = reason
        else:
            # DEFAULT_TAG instead of None
            final_tag  = DEFAULT_TAG
            tag_source = "needs_ai"
    else:
        final_tag  = DEFAULT_TAG
        tag_source = "needs_ai"

    clean_text    = strip_tags_from_text(text)
    text_with_tag = f"[{final_tag}] {clean_text}"

    return {
        "raw_tag":       raw_tag,
        "final_tag":     final_tag,
        "tag_source":    tag_source,
        "text":          clean_text,
        "text_with_tag": text_with_tag,
        "line":          line,
    }


def _process_long_content(
    content: str,
    lang:    str,
) -> list[dict]:
    """
    معالجة Long content:
    كل [tag] + الفقرة التي تليه = وحدة واحدة كاملة.
    لا تقسّم الفقرات إلى جمل.
    """
    parts  = _TAG_SPLIT_RE.split(content.strip())
    result : list[dict] = []
    line   = 1
    i      = 0

    while i < len(parts):
        part = parts[i].strip()

        if not part:
            i += 1
            continue

        if _TAG_SPLIT_RE.fullmatch(part):
            # هذا tag → خذ النص الذي يليه
            raw_tag = part[1:-1].strip()
            text    = (
                parts[i + 1].strip()
                if i + 1 < len(parts) else ""
            )
            i += 2
        else:
            # نص بدون tag
            raw_tag = ""
            text    = part
            i += 1

        if not text or len(text) < MIN_SENTENCE_CHARS:
            continue

        result.append(_process_single_sentence({
            "raw_tag": raw_tag,
            "text":    text,
            "line":    line,
        }))
        line += 1

    log.info(
        "  📝 Long content parsed: %d paragraphs",
        len(result)
    )
    return result


def _process_short_content(
    content: str,
    lang:    str,
) -> list[dict]:
    """
    معالجة Short content:
    يستخدم split_into_tagged_sentences للتقسيم على مستوى الجمل.
    """
    tagged = split_into_tagged_sentences(content)
    result = [
        _process_single_sentence(sent)
        for sent in tagged
        if (sent.get("text") or "").strip()
    ]
    log.info(
        "  📝 Short content parsed: %d sentences",
        len(result)
    )
    return result


def process_tagged_content(
    content:      str,
    lang:         str = "ar",
    content_mode: str = "short",
) -> list[dict]:
    """
    معالجة محتوى يحتوي على tags.

    Short: كل جملة = وحدة منفصلة
    Long:  كل [tag]+فقرة = وحدة واحدة (لا تُقسَّم)

    Args:
        content:      النص الكامل مع tags
        lang:         ar | fr | en
        content_mode: short | long

    Returns:
        list of dicts with: raw_tag, final_tag, tag_source,
        text, text_with_tag, line
    """
    if not content or not content.strip():
        return []

    if content_mode == "long":
        return _process_long_content(content, lang)
    else:
        return _process_short_content(content, lang)


# ═════════════════════════════════════════════════════════════════════════════
# SENTENCE SPLITTING (للنصوص بدون tags)
# ═════════════════════════════════════════════════════════════════════════════

def _split_by_punctuation(text: str, lang: str) -> list[str]:
    """تقسيم بعلامات الترقيم."""
    pattern = _AR_SENTENCE_RE if lang == "ar" else _EN_SENTENCE_RE
    return pattern.split(text.strip())


def _merge_short_sentences(sentences: list[str]) -> list[str]:
    """دمج الجمل القصيرة."""
    cleaned: list[str] = []
    for sent in sentences:
        sent = sent.strip()
        if len(sent) < MIN_SENTENCE_CHARS:
            continue
        if (
            cleaned and
            len(sent.split()) < MIN_SENTENCE_WORDS
        ):
            cleaned[-1] = cleaned[-1] + " " + sent
        else:
            cleaned.append(sent)
    return cleaned


def split_into_sentences(
    text: str,
    lang: str = "en",
) -> list[str]:
    """تقسيم نص إلى جمل (للنصوص بدون tags)."""
    if not text or not text.strip():
        return []
    parts = _split_by_punctuation(text, lang)
    return _merge_short_sentences(parts)


# ═════════════════════════════════════════════════════════════════════════════
# ROW PROCESSING
# ═════════════════════════════════════════════════════════════════════════════

def _process_rows(
    rows:         list,
    col_map:      dict[str, int],
    skip_first:   bool = True,
    file_lang:    str  = "ar",
    content_mode: str  = "short",
) -> list[dict]:
    """
    معالجة صفوف Excel/CSV.

    - يتخطى الصفوف الفارغة
    - يضيف lang و content_mode لكل record
    - يُعطي أرقام تلقائية فريدة
    """
    scripts:      list[dict] = []
    auto_counter: int        = 1
    start_idx                = 1 if skip_first else 0

    for row in rows[start_idx:]:
        if not _row_has_data(row):
            continue

        record = _row_to_dict(list(row), col_map)

        # رقم تلقائي فريد
        if (
            not record.get("number") or
            _normalize(record["number"]) in _INVALID_NUMBERS
        ):
            record["number"] = str(auto_counter)

        auto_counter += 1

        # أضف metadata
        record["lang"]         = file_lang
        record["content_mode"] = content_mode

        # اختر المحتوى المناسب للغة
        lang_content = record.get(
            f"{file_lang}_content", ""
        ).strip()
        generic = record.get("content", "").strip()
        record["content"] = lang_content or generic

        if _is_valid_record(record):
            scripts.append(record)

    return scripts


# ═════════════════════════════════════════════════════════════════════════════
# FILE READING
# ═════════════════════════════════════════════════════════════════════════════

def _print_detected_columns(col_map: dict[str, int]) -> None:
    """طباعة الأعمدة المُكتشَفة."""
    detected = sorted(col_map.keys())
    log.info(
        "  📊 Columns detected: %s",
        ', '.join(detected)
    )


def _read_excel(
    path:         Path,
    file_lang:    str = "ar",
    content_mode: str = "short",
) -> list[dict]:
    """
    قراءة ملف Excel.

    Smart sheet selection: prefers sheets named
    "scripts", "data", "videos", "content", "sheet1".
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl not installed. Run: pip install openpyxl"
        )

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise RuntimeError(
            f"Cannot read Excel file '{path.name}': {e}"
        )

    log.info(
        "  📋 Sheets available: %s", wb.sheetnames
    )

    # Smart sheet selection
    preferred_names = [
        "scripts", "data", "videos", "content", "sheet1"
    ]
    ws = None
    for name in wb.sheetnames:
        if name.lower() in preferred_names:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    try:
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise RuntimeError(
            f"Cannot read rows from '{path.name}': {e}"
        )

    if not rows:
        log.warning(
            "  ⚠️  Excel file is empty: %s", path.name
        )
        return []

    headers = [
        str(c).strip() if c is not None else ""
        for c in rows[0]
    ]

    if not any(headers):
        raise RuntimeError(
            f"Excel file has no valid headers: {path.name}"
        )

    log.info("  📄 Using sheet: '%s'", ws.title)
    col_map = _detect_columns(headers, file_lang)
    _print_detected_columns(col_map)

    return _process_rows(
        rows, col_map,
        skip_first   = True,
        file_lang    = file_lang,
        content_mode = content_mode,
    )


def _read_csv(
    path:         Path,
    file_lang:    str = "ar",
    content_mode: str = "short",
) -> list[dict]:
    """
    قراءة ملف CSV.

    Uses _process_rows (no code duplication).
    """
    try:
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader  = csv.reader(f)
            headers = next(reader, [])

            if not headers:
                log.warning(
                    "  ⚠️  CSV is empty: %s", path.name
                )
                return []

            if not any(h.strip() for h in headers):
                raise RuntimeError(
                    f"CSV has no valid headers: {path.name}"
                )

            col_map  = _detect_columns(headers, file_lang)
            _print_detected_columns(col_map)
            
            # Build rows with headers as first row
            all_rows = [list(headers)] + list(reader)

    except UnicodeDecodeError as e:
        raise RuntimeError(
            f"Cannot decode '{path.name}' as UTF-8: {e}"
        )
    except csv.Error as e:
        raise RuntimeError(
            f"CSV parsing error in '{path.name}': {e}"
        )
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(
            f"Cannot read CSV '{path.name}': {e}"
        )

    # Use shared _process_rows (no code duplication)
    return _process_rows(
        all_rows, col_map,
        skip_first   = True,
        file_lang    = file_lang,
        content_mode = content_mode,
    )


def read_scripts(file_path: str) -> list[dict]:
    """
    قراءة ملف سكريبتات مع اكتشاف اللغة والنوع تلقائياً.

    Args:
        file_path: مسار الملف (.xlsx, .xls, .csv)

    Returns:
        قائمة السكريبتات — كل dict يحتوي:
            number, title, content, lang, content_mode,
            ar_content / fr_content / en_content

    Raises:
        FileNotFoundError: إذا الملف غير موجود
        ValueError:        إذا الصيغة غير مدعومة
        ImportError:       إذا openpyxl غير مثبت
        RuntimeError:      إذا الملف معطوب
    """
    path = Path(file_path).resolve()

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if not path.is_file():
        raise FileNotFoundError(f"Not a file: {path}")

    # اكتشاف lang و content_mode من اسم الملف
    file_lang, content_mode = _detect_file_meta(path)

    ext = path.suffix.lower()

    if ext in EXCEL_EXTENSIONS:
        scripts = _read_excel(path, file_lang, content_mode)
    elif ext in CSV_EXTENSIONS:
        scripts = _read_csv(path, file_lang, content_mode)
    else:
        raise ValueError(
            f"Unsupported format: '{ext}' "
            f"— use .xlsx, .xls, or .csv"
        )

    log.info(
        "  ✅ %d records loaded from '%s' [%s/%s]",
        len(scripts), path.name,
        file_lang.upper(), content_mode.upper()
    )
    return scripts


# ═════════════════════════════════════════════════════════════════════════════
# VALIDATION
# ═════════════════════════════════════════════════════════════════════════════

def _count_tags(content: str) -> int:
    """عد الـ tags في المحتوى."""
    return len(_TAG_RE.findall(content))


def _check_tags_in_content(record: dict) -> list[str]:
    """تحذيرات حول الـ tags."""
    warnings: list[str] = []
    num  = record.get("number", "?")
    lang = record.get("lang", "ar")
    mode = record.get("content_mode", "short").upper()

    content = record.get("content", "").strip()
    if content and _count_tags(content) == 0:
        warnings.append(
            f"  ⚠️  #{num} ({lang.upper()}) [{mode}]: "
            f"no [tags] found — AI will add them"
        )
    return warnings


def _unique_num(num: str, seen: set[str]) -> str:
    """
    ضمان رقم فريد (لا تصادم).

    Tries: num_1, num_2, ... until unique.
    """
    candidate = num
    suffix    = 1
    while candidate in seen:
        candidate = f"{num}_{suffix}"
        suffix   += 1
    return candidate


def validate_scripts(
    scripts: list[dict],
) -> tuple[list[dict], list[str]]:
    """
    التحقق من صحة السكريبتات.

    - يتحقق من وجود محتوى
    - يكتشف الأرقام المكررة ويعيد تسميتها
    - يُنبّه عن غياب الـ tags

    Returns:
        (valid_scripts, error_messages)
    """
    valid:      list[dict] = []
    errors:     list[str]  = []
    seen_nums:  set[str]   = set()

    for script in scripts:
        # تحقق من المحتوى
        if not _has_content(script):
            errors.append(
                f"  ❌ #{script.get('number','?')} "
                f"'{script.get('title','')}': no content found"
            )
            continue

        # تحقق من الأرقام المكررة
        num = str(script.get("number", "")).strip()
        if num in seen_nums:
            new_num = _unique_num(num, seen_nums)
            errors.append(
                f"  ⚠️  #{num} duplicate number "
                f"— renamed to #{new_num}"
            )
            script = {**script, "number": new_num}
            num    = new_num

        seen_nums.add(num)

        # تحذيرات الـ tags
        errors.extend(_check_tags_in_content(script))
        valid.append(script)

    return valid, errors


# ═════════════════════════════════════════════════════════════════════════════
# SUMMARY DISPLAY
# ═════════════════════════════════════════════════════════════════════════════

def _get_preview_text(record: dict) -> str:
    """جلب نص المعاينة."""
    for key in ("content",) + ALL_CONTENT_KEYS:
        content = record.get(key, "")
        if content:
            return content[:PREVIEW_LENGTH].replace("\n", " ")
    return ""


def _build_language_flags(record: dict) -> list[str]:
    """بناء أعلام اللغة."""
    flags = []
    lang  = record.get("lang", "ar")

    content = record.get("content", "").strip()
    if content:
        tags_count = _count_tags(content)
        mode       = record.get("content_mode", "short").upper()
        flags.append(
            f"🌐 {lang.upper()} ({tags_count} tags) [{mode}]"
        )

    return flags


def _print_script_entry(record: dict) -> None:
    """طباعة معلومات سكريبت واحد."""
    num     = record.get("number", "?")
    title   = str(record.get("title", ""))[:TITLE_MAX_LENGTH]
    preview = _get_preview_text(record)
    flags   = _build_language_flags(record)

    log.info("  #%s  %s", str(num).rjust(3), title)
    if preview:
        log.info("       %s...", preview)
    if flags:
        log.info("       %s", ' | '.join(flags))


def print_scripts_summary(scripts: list[dict]) -> None:
    """طباعة ملخص السكريبتات المحملة."""
    sep = "═" * SUMMARY_WIDTH

    log.info("\n%s", sep)
    log.info("  📋  %d scripts loaded", len(scripts))
    log.info("%s", sep)

    for script in scripts:
        _print_script_entry(script)

    log.info("%s\n", sep)
